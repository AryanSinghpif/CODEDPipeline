"""
Build a single Excel workbook from extracted tables:
one sheet per table (heading row = title from the PDF),
plus an Index sheet with hyperlinks to every table.
"""

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="C8102E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
LINK_FONT = Font(color="0563C1", underline="single")


def _sheet_name(title, table_id, used):
    """Excel sheet names: <=31 chars, unique, no \\/?*[]: characters."""

    m = re.search(r"Table\s+(\d+\.\d+)", str(title))
    base = f"Table_{m.group(1).replace('.', '_')}" if m else f"table_{table_id}"

    base = re.sub(r"[\\/?*\[\]:]", "_", base)[:28]

    name, n = base, 2
    while name in used:
        name = f"{base}_{n}"
        n += 1

    used.add(name)
    return name


def _maybe_number(value):
    """Store numerics as numbers so Excel can compute on them."""

    text = str(value).strip().replace(",", "")
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d*\.\d+", text):
        return float(text)
    return value


def build_workbook(table_dfs, catalog):
    """
    table_dfs: {table_id: DataFrame}
    catalog:   list of metadata dicts with table_id, table_name, page
    Returns BytesIO of the .xlsx file.
    """

    meta = {m["table_id"]: m for m in catalog}

    wb = Workbook()
    index_ws = wb.active
    index_ws.title = "Index"

    used_names = set()
    index_rows = []

    for tid, df in table_dfs.items():

        info = meta.get(tid, {})
        title = info.get("table_name", f"table_{tid}")
        page = info.get("page", "")

        sheet = _sheet_name(title, tid, used_names)
        ws = wb.create_sheet(sheet)

        n_cols = max(df.shape[1], 1)

        # title row (from the PDF), merged across the table width
        ws.cell(row=1, column=1, value=str(title)).font = TITLE_FONT
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=n_cols
        )

        # column header row
        for c, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=3, column=c, value=str(col))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        # data
        for r, row in enumerate(df.itertuples(index=False), start=4):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=_maybe_number(value))

        # column widths
        for c, col in enumerate(df.columns, start=1):
            width = max(
                len(str(col)),
                *(len(str(v)) for v in df.iloc[:, c - 1].head(50)),
                8,
            )
            ws.column_dimensions[get_column_letter(c)].width = min(width + 2, 45)

        ws.freeze_panes = "A4"
        index_rows.append((tid, title, page, sheet))

    # index sheet
    for c, h in enumerate(["Table ID", "Table Name", "PDF Page", "Sheet"], 1):
        cell = index_ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for r, (tid, title, page, sheet) in enumerate(index_rows, start=2):
        index_ws.cell(row=r, column=1, value=tid)
        link = index_ws.cell(row=r, column=2, value=str(title))
        link.hyperlink = f"#'{sheet}'!A1"
        link.font = LINK_FONT
        index_ws.cell(row=r, column=3, value=page)
        index_ws.cell(row=r, column=4, value=sheet)

    index_ws.column_dimensions["A"].width = 10
    index_ws.column_dimensions["B"].width = 70
    index_ws.column_dimensions["C"].width = 10
    index_ws.column_dimensions["D"].width = 18
    index_ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
